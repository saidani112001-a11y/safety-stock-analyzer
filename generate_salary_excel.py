import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_salary_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaire Maroc"

    header_fill = PatternFill(start_color="FFEDEDED", end_color="FFEDEDED", fill_type="solid")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    widths = {
        1: 30, 2: 20, 3: 20
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 1
    ws.cell(row=row, column=1, value="Paramètres").font = Font(bold=True, size=13)
    row += 1

    # Inputs
    inputs = [
        ("Jours travaillés (mois)", 26),
        ("Salaire journalier de base", 0),
        ("Heures supp 125% (nombre d'heures)", 0),
        ("Heures supp 150% (nombre d'heures)", 0),
        ("Taux horaire normal (MAD/heure)", 0),
        ("Cotisation retraite complémentaire (CIMR) %", 0.0),
        ("Personnes à charge (réduction IR)", 0),
    ]

    ws.append(["Libellé", "Valeur", "Note"])
    for col in range(1, 4):
        c = ws.cell(row=row, column=col)
        c.font = bold
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1

    start_input_row = row
    for label, default in inputs:
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=default).border = border
        ws.cell(row=row, column=3, value="Entrée utilisateur").border = border
        row += 1

    # Derived calculations and breakdown
    row += 1
    ws.cell(row=row, column=1, value="Calculs").font = Font(bold=True, size=13)
    row += 1

    calc_start = row

    # Names for input cells for clarity
    # Map input indexes
    jours_cell = f"$B${start_input_row}"
    sj_cell = f"$B${start_input_row + 1}"
    hs125_cell = f"$B${start_input_row + 2}"
    hs150_cell = f"$B${start_input_row + 3}"
    taux_h_cell = f"$B${start_input_row + 4}"
    cimr_cell = f"$B${start_input_row + 5}"
    charges_cell = f"$B${start_input_row + 6}"

    # Brut components
    ws.append(["Salaire de base mensuel", None, "= Salaire journalier × Jours"])
    ws.cell(row=row, column=2).value = f"={sj_cell}*{jours_cell}"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["Heures supp 125% (montant)", None, "= heures × taux × 125%"])
    ws.cell(row=row, column=2).value = f"={hs125_cell}*{taux_h_cell}*1.25"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["Heures supp 150% (montant)", None, "= heures × taux × 150%"])
    ws.cell(row=row, column=2).value = f"={hs150_cell}*{taux_h_cell}*1.5"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["Salaire brut mensuel", None, "= base + HS125 + HS150"])
    base_cell = f"$B${calc_start}"
    hs125_amt_cell = f"$B${calc_start + 1}"
    hs150_amt_cell = f"$B${calc_start + 2}"
    brut_cell = f"$B${row}"
    ws.cell(row=row, column=2).value = f"={base_cell}+{hs125_amt_cell}+{hs150_amt_cell}"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Social contributions (employee share)
    ws.append(["CNSS (4,48% plaf. 6 000)", None, "min(6 000 ; Brut) × 4,48%"])
    cnss_row = row
    ws.cell(row=row, column=2).value = f"=MIN(6000,{brut_cell})*0.0448"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["AMO (2,26% du Brut)", None, "Brut × 2,26%"])
    amo_row = row
    ws.cell(row=row, column=2).value = f"={brut_cell}*0.0226"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["CIMR (si applicable)", None, "Brut × % CIMR"])
    cimr_row = row
    ws.cell(row=row, column=2).value = f"={brut_cell}*({cimr_cell}/100)"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["Cotisations sociales (total)", None, "CNSS + AMO + CIMR"])
    cotis_row = row
    ws.cell(row=row, column=2).value = f"=$B${cnss_row}+$B${amo_row}+$B${cimr_row}"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Abattement frais pro (sur Brut - Cotisations)
    ws.append(["Abattement frais professionnels (20% plaf. 2 500)", None, "min(20%×(Brut − Cotisations) ; 2500)"])
    abatt_row = row
    ws.cell(row=row, column=2).value = f"=MIN(({brut_cell}-$B${cotis_row})*0.2,2500)"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Net imposable
    ws.append(["Salaire net imposable (SNI)", None, "Brut − Cotisations − Abattement"])
    sni_row = row
    sni_cell = f"$B${sni_row}"
    ws.cell(row=row, column=2).value = f"={brut_cell}-$B${cotis_row}-$B${abatt_row}"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # IR calculation by monthly brackets
    # Brackets (monthly):
    # 0 – 2,500: 0%
    # 2,501 – 4,166: 10%
    # 4,167 – 5,000: 20%
    # 5,001 – 6,666: 30%
    # 6,667 – 15,000: 34%
    # > 15,000: 38%

    # Helper: tax per bracket via Excel formula using MAX/MIN
    def tranche(lower: int, upper: int | None, rate: float) -> str:
        if upper is None:
            return f"=MAX(0,{sni_cell}-{lower})*{rate}"
        return f"=MAX(0,MIN({sni_cell},{upper})-{lower})*{rate}"

    ws.append(["IR tranche 0%", None, "jusqu'à 2 500"])
    ws.cell(row=row, column=2).value = tranche(0, 2500, 0.0)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR tranche 10%", None, "2 501 à 4 166"])
    ws.cell(row=row, column=2).value = tranche(2500, 4166, 0.10)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR tranche 20%", None, "4 167 à 5 000"])
    ws.cell(row=row, column=2).value = tranche(4166, 5000, 0.20)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR tranche 30%", None, "5 001 à 6 666"])
    ws.cell(row=row, column=2).value = tranche(5000, 6666, 0.30)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR tranche 34%", None, "6 667 à 15 000"])
    ws.cell(row=row, column=2).value = tranche(6666, 15000, 0.34)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR tranche 38%", None, "> 15 000"])
    ws.cell(row=row, column=2).value = tranche(15000, None, 0.38)
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    first_ir_row = sni_row + 1
    last_ir_row = row - 1

    ws.append(["IR brut (somme des tranches)", None, "Somme IR tranches"])
    ir_brut_row = row
    ir_brut_cell = f"$B${ir_brut_row}"
    ws.cell(row=row, column=2).value = f"=SUM($B${first_ir_row}:$B${last_ir_row})"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Réduction charges de famille: 30 MAD par personne, max 180
    ws.append(["Réduction IR (charges de famille)", None, "30 MAD × nb, max 180"])
    reduc_row = row
    ws.cell(row=row, column=2).value = f"=MIN({charges_cell}*30,180)"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    ws.append(["IR net", None, "max(IR brut − réduction, 0)"])
    ir_net_row = row
    ir_net_cell = f"$B${ir_net_row}"
    ws.cell(row=row, column=2).value = f"=MAX({ir_brut_cell}-$B${reduc_row},0)"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Salaire net à payer
    ws.append(["Salaire net à payer", None, "Brut − Cotisations − IR net"])
    ws.cell(row=row, column=2).value = f"={brut_cell}-$B${cotis_row}-{ir_net_cell}"
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)

    # Styling headings
    for r in [1, calc_start - 1]:
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = header_fill

    # Freeze panes below header row
    ws.freeze_panes = ws["A3"]

    return wb


def main() -> None:
    wb = create_salary_workbook()
    output_path = "salary_calculator_maroc.xlsx"
    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    main()

